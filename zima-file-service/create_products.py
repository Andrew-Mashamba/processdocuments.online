#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import os

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

# Headers
headers = ["Product", "Price"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

# Product data
products = [
    ["iPhone", "$999"],
    ["MacBook", "$1999"],
    ["iPad", "$799"]
]

# Add data rows
for row_idx, product in enumerate(products, 2):
    for col_idx, value in enumerate(product, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Style the data
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply borders to all cells
for row in ws.iter_rows(min_row=1, max_row=len(products)+1, min_col=1, max_col=len(headers)):
    for cell in row:
        cell.border = thin_border

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the file
output_path = "/Volumes/DATA/QWEN/zima-file-service/generated_files/products_final.xlsx"
wb.save(output_path)

print(f"Excel file created successfully: {output_path}")
print(f"File contains {len(products)} products with headers: {', '.join(headers)}")

# Verify the file was created and is readable
try:
    test_wb = openpyxl.load_workbook(output_path)
    test_ws = test_wb.active
    print("\nFile verification successful!")
    print("Contents:")
    for row in test_ws.iter_rows(values_only=True):
        print(f"  {row}")
except Exception as e:
    print(f"File verification failed: {e}")