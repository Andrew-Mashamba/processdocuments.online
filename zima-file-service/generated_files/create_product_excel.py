#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Product Data"

# Headers
headers = ["Product ID", "Product Name", "Category", "Price", "Stock Quantity", "Supplier", "Last Updated"]

# Product data rows
rows = [
    ["P001", "Laptop Computer", "Electronics", "$1299.99", 45, "TechCorp", "2026-01-15"],
    ["P002", "Office Chair", "Furniture", "$249.99", 23, "ComfortSeating", "2026-01-20"],
    ["P003", "Coffee Maker", "Appliances", "$89.99", 67, "BrewMaster", "2026-01-18"],
    ["P004", "Wireless Mouse", "Electronics", "$29.99", 134, "TechCorp", "2026-01-22"],
    ["P005", "Desk Lamp", "Furniture", "$45.99", 78, "LightWorks", "2026-01-16"],
    ["P006", "Water Bottle", "Accessories", "$19.99", 89, "HydroLife", "2026-01-25"],
    ["P007", "Keyboard", "Electronics", "$79.99", 56, "TechCorp", "2026-01-21"],
    ["P008", "Plant Pot", "Home & Garden", "$12.99", 145, "GreenThumb", "2026-01-19"],
    ["P009", "Notebook Set", "Office Supplies", "$15.99", 203, "PaperPlus", "2026-01-24"],
    ["P010", "Phone Charger", "Electronics", "$24.99", 98, "PowerTech", "2026-01-23"]
]

# Add headers to the first row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data rows
for row_num, row_data in enumerate(rows, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)
        # Alternate row colors
        if row_num % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save the file
output_path = "/Volumes/DATA/QWEN/zima-file-service/generated_files/product_data.xlsx"
wb.save(output_path)
print(f"Excel file created successfully: {output_path}")